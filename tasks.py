from robocorp import browser
from robocorp.tasks import task
from RPA.Excel.Files import Files as Excel
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from robocorp import windows
import os 
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from pathlib import Path
import os
import requests
import time
from functions import read_excel_preingreso, read_excel_envio, compare_results

FILE_NAME = "challenge.xlsx"
OUTPUT_DIR = Path(os.environ.get('ROBOT_ARTIFACTS'))
EXCEL_URL = f"https://rpachallenge.com/assets/downloadFiles/{FILE_NAME}"
EXCEL_FILE_NAME = "//DESKTOP-E3HRBD4/Compartidas/Itau_cpago_gtia_estatal/ArchivoInput/Julio_2024.xlsx"
PATH_BOT = "C:\Applications\RPA 01 - INGRESO DE ESCRITOS DE PAGO GARANTIA ESTATAL EN PODER JUDICIAL"
ESCRITOS_DCP_PATH = "\\\\DESKTOP-E3HRBD4\\Compartidas\\Itau_cpago_gtia_estatal\\Escritos Da Cuenta de Pago"

ROLES_DONE = []

#@task
def RPA_01_Ingreso():
    """
    RPA 01 - Ingreso de escritos de pago garantia estatal en poder judicial.
    
    """
    browser.configure(
        browser_engine="chrome",
        screenshot="only-on-failure",
        headless=False,
    )
    try:
        page = browser.goto("https://ojv.pjud.cl/kpitec-ojv-web/views/login.html")
        page.get_by_role("heading", name="Ingreso de demandas y escritos").wait_for()
        page.get_by_role("button", name="Clave Poder Judicial").first.click()
        page.get_by_role("textbox", name="Ej: 12345678 (Rut sin guión").click()
        page.get_by_role("textbox", name="Ej: 12345678 (Rut sin guión").fill("11346197")
        page.locator("#inputPassword2C").click()
        page.locator("#inputPassword2C").fill("Talaveras1551+")
        page.get_by_role("button", name="Ingresar").click()
        page.locator("#roles-modal").get_by_text("Seleccione Perfil").wait_for()
        page.get_by_text("11.346.197-7").click()

        read_excel_preingreso(EXCEL_FILE_NAME, ESCRITOS_DCP_PATH)
        
    finally:
        # Place for teardown and cleanups
        # Playwright handles browser closing
        print('Done')
    
@task
def RPA_01_Envio():
    """
    RPA 01 - Ingreso de escritos de pago garantia estatal en poder judicial.
    
    """
    browser.configure(
        browser_engine="chromium",
        screenshot="only-on-failure",
        headless=False,
    )
    try:
        page = browser.goto("https://ojv.pjud.cl/kpitec-ojv-web/views/login.html")
        page.get_by_role("heading", name="Ingreso de demandas y escritos").wait_for()
        page.get_by_role("button", name="Clave Poder Judicial").first.click()
        page.get_by_role("textbox", name="Ej: 12345678 (Rut sin guión").click()
        page.get_by_role("textbox", name="Ej: 12345678 (Rut sin guión").fill("11346197")
        page.locator("#inputPassword2C").click()
        page.locator("#inputPassword2C").fill("Talaveras1551+")
        page.get_by_role("button", name="Ingresar").click()
        page.locator("#roles-modal").get_by_text("Seleccione Perfil").wait_for()
        page.get_by_text("11.346.197-7").click()

        read_excel_envio(EXCEL_FILE_NAME)

    finally:
        # Place for teardown and cleanups
        # Playwright handles browser closing
        print('Done')

#@task
def RPA_01_Compare_results():
    """
    RPA 01 - Descarga del Excel del poder judicial y comparacion con el Excel de ingreso del bot.
    
    """
    try:
        chrome_options = Options()

        download_path = PATH_BOT + '\processing'
        # For windows:
        prefs = {'download.default_directory': download_path}

        chrome_options.add_argument('--start-maximized')
        #chrome_options.add_argument("--headless=old")
        chrome_options.add_experimental_option('prefs', prefs)       

        driver = webdriver.Chrome(options=chrome_options)
        #driver = webdriver.Chrome()
        driver.get("https://ojv.pjud.cl/kpitec-ojv-web/views/login.html")

        driver.find_element(By.ID, value='link2C').click()
        driver.find_element(By.ID, value='inputRut2C').send_keys('11346197')
        driver.find_element(By.ID, value='inputPassword2C').send_keys('Talaveras1551+')
        driver.find_element(By.CLASS_NAME, value='btn-ingreso-pjud').click()

        time.sleep(2)
        driver.find_element(By.CLASS_NAME, value='card-text').click()
        time.sleep(2)
        driver.get("https://ojv.pjud.cl/kpitec-ojv-web/index#bandeja/escritos")

        time.sleep(5)
        driver.find_element(By.XPATH, "//select[contains(.,'Seleccione Competencia')]/option[text()='Civil']").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//a[contains(.,'Escritos Enviados')]").click()
        time.sleep(5)
        driver.find_element(By.XPATH, "//button[contains(.,'Exportar Excel')]").click()
        time.sleep(15)

        #app = windows.find_window('regex:.*Oficina Judicial Virtual - Portal*')
        #time.sleep(3)
        #controlFile_path = PATH_BOT + '\input\ArchivoControl.xlsx'
        #app.find('control:"EditControl" and name:"File name:"').set_value(controlFile_path)
        #time.sleep(3)
        #app.find('class:"Button" and name:"Save"').click();
        #time.sleep(2)

        time.sleep(15) # Let the user actually see something!
        
        ## Archivo descargado
        
        ## Busco el archivo
        path = PATH_BOT + "\processing"
        outputPath = PATH_BOT + "\output"
        
        dir_list = os.listdir(path) 
        print("Files and directories in '", path, "' :") 
        print(dir_list) 
        file = dir_list[0]
        path = os.path.join(path, file)

        informePJUD = openpyxl.load_workbook(path)
        escritos_enviados_sheet = informePJUD["Escritos Enviados"]

        get_roles_done_by_excel()
  
        # Iterate over the columns in the sheet 
        for column in escritos_enviados_sheet.iter_cols(): 
            # Get the value of the first cell in the column  
            # (the cell with the column name) 
            
            column_name = column[3].value 
            # Check if the column is the "Name" column 
            if column_name == "RIT": 
                # Iterate over the cells in the column 
                for i, cell in enumerate(column): 
                    # Skip the first cell (the cell with the column name) 
                    if i < 4: 
                        continue
                    # Add the value of the cell to the list 
                    #rit_list.append(cell.value) 
                    print(cell.value, 'pjud')
                    if cell.value in ROLES_DONE :
                        print(cell.value, 'into roles_done')
                        escritos_enviados_sheet.cell(cell.row,cell.column, cell.value).fill = PatternFill(start_color='008000', end_color='008000', fill_type="solid")
  
        # Print the list of rit 
        #print(rit_list, len(rit_list)) 
        outputPath = os.path.join(outputPath, 'RESULTADO.xlsx')
        informePJUD.save(outputPath)        
        
        ## elimino el archivo
         
        #os.remove(path) 

    finally:
        # Place for teardown and cleanups
        # Playwright handles browser closing
        #browser.close();
        #driver.quit()
        print('Done')

def get_roles_done_by_excel():
    """
    RPA 01 - Review column ENVIO into an Excel file and set the value of ROLES_DONE variable
    
    """
    try:
        ## Busco el archivo
        informe_bot = openpyxl.load_workbook(EXCEL_FILE_NAME)
        informe_bot_sheet = informe_bot["Hoja1"]

        rit_list = []
        
        lista = list(informe_bot_sheet.iter_cols());
        envio_index = None
        rol_index = None

        for index in range(len(lista)):
            if (lista[index][0].value == 'ENVIO') :
                envio_index = index
            elif (lista[index][0].value == 'ROL') :
                rol_index = index
        
        for index in range(informe_bot_sheet.max_row):    
            if (lista[envio_index][index].value == 'OK') :
                rit_list.append(lista[rol_index][index].value)
            elif (lista[envio_index][index].value == None and lista[rol_index][index].value == None) :
                break

        global ROLES_DONE
        ROLES_DONE = list(map(lambda item : "C-"+str(item) ,rit_list))
        return ROLES_DONE

    finally:
        informe_bot.close